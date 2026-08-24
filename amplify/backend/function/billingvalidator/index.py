import json
import boto3
import boto3.dynamodb.conditions
from botocore.exceptions import ClientError
import pandas as pd
import numpy as np
from io import StringIO, BytesIO
from datetime import datetime, timedelta
import urllib.parse
import os
import traceback
import re

s3 = boto3.client('s3')
dynamodb = boto3.resource('dynamodb')

CONFIG_BUCKET = "billing-config-amh"
OUTPUT_BUCKET = "billing-output-amh"


def generate_presigned_urls(output_files, expires_in=86400):
    """Generate presigned URLs for all output files"""
    presigned_urls = []
    for file_info in output_files:
        try:
            url = s3.generate_presigned_url(
                'get_object',
                Params={
                    'Bucket': OUTPUT_BUCKET,
                    'Key': file_info['s3Key'],
                    'ResponseContentDisposition': f'attachment; filename="{file_info["fileName"]}"'
                },
                ExpiresIn=expires_in
            )
            presigned_urls.append({
                'fileName': file_info['fileName'],
                'url': url,
                's3Key': file_info['s3Key'],
                'type': file_info['type'],
            })
            print(f"✅ Generated presigned URL for {file_info['fileName']}")
        except Exception as e:
            print(f"❌ Failed to generate presigned URL for {file_info['fileName']}: {str(e)}")
    return presigned_urls


def eom_date(date_obj):
    if pd.isna(date_obj):
        return None
    next_month = date_obj.replace(day=28) + timedelta(days=4)
    return next_month - timedelta(days=next_month.day)


def calculate_tax_split(gross_amount, tax_amount):
    net_20 = tax_amount * 5 if tax_amount > 0 else 0
    net_0 = gross_amount - tax_amount - net_20
    if abs(net_0) < 0.20:
        net_20 += net_0
        net_0 = 0
    # Round to 2 decimal places to avoid floating point errors
    return round(net_0, 2), round(net_20, 2), round(tax_amount, 2)


def get_gl_code(dimension_type):
    gl_codes = {
        "Education": 4000,
        "Care": 4001,
        "Leisure": 4002,
        "Contractor": 4003
    }
    return gl_codes.get(dimension_type, 4000)


def get_client_name(class_id, dimensions_df):
    client_row = dimensions_df[dimensions_df.iloc[:, 0] == class_id]
    if not client_row.empty:
        return client_row.iloc[0, 1]
    return "Unknown_Client"


def clean_client_label(label):
    """
    Strip any trailing ' - Unknown' or standalone 'Unknown' fragments from a
    client label so they never appear on the client-facing summary sheet, tab,
    or filename.
    e.g. "St Paul's Girls' School - Tuck Account - Unknown"
      -> "St Paul's Girls' School - Tuck Account"
    """
    label = str(label).strip()
    label = re.sub(r'\s*-\s*Unknown\s*$', '', label, flags=re.IGNORECASE).strip()
    if label.lower() == 'unknown':
        label = ''
    return label


def load_from_s3(bucket, key, sheet_name=None):
    response = s3.get_object(Bucket=bucket, Key=key)
    file_content = response['Body'].read()
    if key.endswith('.xlsx'):
        if sheet_name:
            return pd.read_excel(BytesIO(file_content), sheet_name=sheet_name, engine="openpyxl")
        else:
            return pd.read_excel(BytesIO(file_content), engine="openpyxl")
    else:
        return pd.read_csv(BytesIO(file_content))


def upload_to_s3(df, bucket, key, file_type='excel'):
    if file_type == 'excel':
        buffer = BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df.to_excel(writer, index=False)
        buffer.seek(0)
        s3.put_object(Bucket=bucket, Key=key, Body=buffer.getvalue())
    else:
        csv_buffer = StringIO()
        df.to_csv(csv_buffer, index=False)
        s3.put_object(Bucket=bucket, Key=key, Body=csv_buffer.getvalue())


def upload_summary_to_s3(df, bucket, key, totals_block=None, sheet_name="Summary", client_label=""):
    """Upload summary with:
    - A1:       Client name + period (large bold) — client-facing label
    - D1/E1/F1/G1: SUPPLIER TOTAL label + Net/VAT/Gross values
    - D2/E2/F2/G2: MARK UP TOTAL
    - D3/E3/F3/G3: GRAND TOTAL
    - Row 4:    Column headers (blue)
    - Row 5+:   Data rows; Subtotal rows are bold + grey
    - Sheet tab named after the client (max 31 chars, Excel-safe)
    - Columns auto-sized to content so numbers never show as ####

    A1 sits in col A; totals block uses cols D-G on the same rows — no clash,
    no row offset needed.
    """
    from openpyxl.styles import PatternFill, Font
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    buffer = BytesIO()
    wb = Workbook()
    ws = wb.active

    # Name the sheet tab after the client.
    # Excel tab names max 31 chars; strip characters Excel disallows.
    safe_tab = re.sub(r'[\\/?*\[\]:]', '', str(sheet_name))[:31]
    ws.title = safe_tab

    columns = ["Supplier", "Invoice Number", "Date", "Center", "Net (£)", "VAT (£)", "Gross (£)"]
    financial_columns = {"Net (£)", "VAT (£)", "Gross (£)"}
    currency_fmt = '£#,##0.00'

    title_font    = Font(bold=True, size=14)
    bold_large    = Font(bold=True, size=12)
    bold_normal   = Font(bold=True)
    header_fill   = PatternFill(start_color="0C6FA9", end_color="0C6FA9", fill_type="solid")
    header_font   = Font(color="FFFFFF", bold=True)
    subtotal_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    # --- A1: Client name + period ---
    # Occupies col A only; totals block uses cols D-G on the same rows so there
    # is no clash and no row offset is needed.
    if client_label:
        cell = ws.cell(row=1, column=1, value=client_label)
        cell.font = title_font

    # --- Rows 1-3: SUPPLIER TOTAL / MARK UP TOTAL / GRAND TOTAL in cols D-G ---
    summary_rows = [
        ("SUPPLIER TOTAL:", totals_block["supplier_total"] if totals_block else {}),
        ("MARK UP TOTAL:",  totals_block["markup_total"]   if totals_block else {}),
        ("GRAND TOTAL:",    totals_block["grand_total"]    if totals_block else {}),
    ]
    for excel_row, (label, vals) in enumerate(summary_rows, start=1):
        cell = ws.cell(row=excel_row, column=4, value=label)
        cell.font = bold_large
        for col_name, col_idx in [("Net (£)", 5), ("VAT (£)", 6), ("Gross (£)", 7)]:
            val = vals.get(col_name, "")
            cell = ws.cell(row=excel_row, column=col_idx, value=val)
            cell.font = bold_large
            if isinstance(val, (int, float)):
                cell.number_format = currency_fmt

    # --- Row 4: column headers ---
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=4, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font

    # --- Row 5+: data rows ---
    for row_idx, (_, row) in enumerate(df.iterrows(), start=5):
        is_subtotal = "Subtotal" in str(row.get("Supplier", ""))
        for col_idx, col_name in enumerate(columns, start=1):
            value = row.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if is_subtotal:
                cell.font = bold_normal
                cell.fill = subtotal_fill
            if col_name in financial_columns and isinstance(value, (int, float)):
                cell.number_format = currency_fmt

    # --- Auto-size all columns ---
    for col_idx, col_name in enumerate(columns, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = len(col_name)
        for excel_row in range(1, ws.max_row + 1):
            cell_val = ws.cell(row=excel_row, column=col_idx).value
            if cell_val is not None:
                if isinstance(cell_val, (int, float)):
                    display = f"£{cell_val:,.2f}"
                else:
                    display = str(cell_val)
                max_len = max(max_len, len(display))
        ws.column_dimensions[col_letter].width = max_len + 2

    wb.save(buffer)
    buffer.seek(0)
    s3.put_object(Bucket=bucket, Key=key, Body=buffer.getvalue())


def update_file_metadata(file_id, updates):
    try:
        table_name = os.environ.get('FILE_METADATA_TABLE')
        if not table_name:
            print("❌ FILE_METADATA_TABLE environment variable not set")
            return None

        print(f"📝 Updating DynamoDB for file {file_id} with: {updates}")
        table = dynamodb.Table(table_name)

        update_expression = "SET "
        expression_attribute_values = {}
        expression_attribute_names = {}

        for key, value in updates.items():
            update_expression += f"#{key} = :{key}, "
            expression_attribute_names[f"#{key}"] = key
            expression_attribute_values[f":{key}"] = value

        update_expression = update_expression[:-2]
        update_expression += ", #updatedAt = :updatedAt"
        expression_attribute_names["#updatedAt"] = "updatedAt"
        expression_attribute_values[":updatedAt"] = datetime.now().isoformat()

        response = table.update_item(
            Key={'id': file_id},
            UpdateExpression=update_expression,
            ExpressionAttributeNames=expression_attribute_names,
            ExpressionAttributeValues=expression_attribute_values,
            ReturnValues="UPDATED_NEW"
        )
        print(f"✅ Successfully updated file metadata for {file_id}")
        return response

    except Exception as e:
        print(f"❌ Error updating DynamoDB: {str(e)}")
        print(f"Stack trace: {traceback.format_exc()}")
        return None


def get_file_id_from_s3_metadata(bucket, key):
    try:
        response = s3.head_object(Bucket=bucket, Key=key)
        metadata = response.get('Metadata', {})
        file_id = metadata.get('fileid') or metadata.get('fileId')
        if file_id:
            print(f"🔍 Found file ID from S3 metadata: {file_id}")
            return file_id
        print("❌ No file ID found in S3 metadata")
        return None
    except Exception as e:
        print(f"❌ Error reading S3 metadata: {str(e)}")
        return None


def find_file_id_by_s3_key(s3_key):
    try:
        table_name = os.environ.get('FILE_METADATA_TABLE')
        if not table_name:
            print("❌ FILE_METADATA_TABLE environment variable not set")
            return None

        print(f"🔍 Searching DynamoDB for S3 key: {s3_key}")
        table = dynamodb.Table(table_name)

        response = table.scan(
            FilterExpression=boto3.dynamodb.conditions.Attr('s3Key').eq(s3_key)
        )
        if response['Items'] and len(response['Items']) > 0:
            file_id = response['Items'][0]['id']
            print(f"✅ Found file by exact s3Key match: {file_id}")
            return file_id

        if not s3_key.startswith('public/'):
            s3_key_with_public = 'public/' + s3_key
            print(f"🔍 Also trying WITH 'public/' prefix: {s3_key_with_public}")
            response = table.scan(
                FilterExpression=boto3.dynamodb.conditions.Attr('s3Key').eq(s3_key_with_public)
            )
            if response['Items'] and len(response['Items']) > 0:
                file_id = response['Items'][0]['id']
                print(f"✅ Found file WITH 'public/' prefix: {file_id}")
                return file_id

        elif s3_key.startswith('public/'):
            s3_key_without_public = s3_key[7:]
            print(f"🔍 Also trying WITHOUT 'public/' prefix: {s3_key_without_public}")
            response = table.scan(
                FilterExpression=boto3.dynamodb.conditions.Attr('s3Key').eq(s3_key_without_public)
            )
            if response['Items'] and len(response['Items']) > 0:
                file_id = response['Items'][0]['id']
                print(f"✅ Found file WITHOUT 'public/' prefix: {file_id}")
                return file_id

        print("❌ No exact s3Key match found, trying filename match...")
        filename = s3_key.split('/')[-1]
        response = table.scan(
            FilterExpression=boto3.dynamodb.conditions.Attr('fileName').eq(filename)
        )
        if response['Items'] and len(response['Items']) > 0:
            file_id = response['Items'][0]['id']
            print(f"✅ Found file by filename match: {file_id}")
            return file_id

        print(f"❌ No file found for S3 key: {s3_key}")
        print("📊 Available files in DynamoDB:")
        all_files = table.scan()
        for item in all_files['Items']:
            print(f"   - {item.get('fileName', 'N/A')} -> s3Key: {item.get('s3Key', 'N/A')}")
        return None

    except Exception as e:
        print(f"❌ Error finding file ID: {str(e)}")
        return None


def validate_units(report_df, units_df):
    report_df["Unit Code"] = report_df.iloc[:, 6].str.split().str[0]
    billable_codes = set(units_df.iloc[:, 0].astype(str))
    report_df = report_df[report_df["Unit Code"].isin(billable_codes)]
    missing_units = set(report_df["Unit Code"]) - billable_codes
    if missing_units:
        raise ValueError(f"Missing units: {missing_units}")
    return report_df


def create_data_sheet(report_df, units_df, dimensions_df):
    data_df = pd.DataFrame({
        "Unit Code": report_df["Unit Code"],
        "Class ID": report_df.iloc[:, 0],
        "Customer ID": report_df.iloc[:, 1],
        "Supplier ID": report_df.iloc[:, 2],
        "Supplier Name": report_df.iloc[:, 3],
        "AP Purchase Invoice Number": report_df.iloc[:, 4],
        "Date": pd.to_datetime(report_df.iloc[:, 5], format='%d/%m/%Y', errors="coerce"),
        "Description": report_df.iloc[:, 6],
        "Mark Up": report_df.iloc[:, 7],
        "Mark Up Check": report_df["Unit Code"] + "-" + report_df.iloc[:, 2].astype(str),
        "Transaction Amount": report_df.iloc[:, 8],
        "Transaction Tax": report_df.iloc[:, 9],
        "Total Amount": report_df.iloc[:, 10],
        "Period Code": report_df.iloc[:, 11]
    })
    # Fill NaN values in all columns except Date (preserve datetime)
    data_df = data_df.fillna({
        col: "Unknown" for col in data_df.columns if col != "Date"
    })
    data_df = add_dimension_columns(data_df, units_df, dimensions_df)
    data_df.sort_values(by=["Supplier ID", "Class ID"], ascending=False, inplace=True)
    return data_df


def add_dimension_columns(data_df, units_df, dimensions_df):
    units_lookup = units_df.iloc[:, [0, 13, 18]]
    units_lookup.columns = ["Unit Code", "Invoice_Grouping_Type", "Email_Grouping_Type"]
    data_df = data_df.merge(units_lookup, on="Unit Code", how="left")

    for index, row in data_df.iterrows():
        grouping_type = row["Invoice_Grouping_Type"]
        class_id = row["Class ID"]
        dim_row = dimensions_df[dimensions_df.iloc[:, 0] == class_id]

        if dim_row.empty:
            print(f"⚠️ No dimension row found for Class ID '{class_id}' "
                  f"(Unit Code: {row['Unit Code']}). "
                  f"Falling back to Class ID as Invoice Grouping Code.")
            data_df.at[index, "Invoice Grouping Code"] = class_id
            data_df.at[index, "Invoice Grouping Name"] = class_id
            data_df.at[index, "Invoice Customer ID"]   = class_id
            data_df.at[index, "D4 Code"]               = class_id
            data_df.at[index, "D4 Name"]               = class_id
            data_df.at[index, "Email Code"]            = class_id
            data_df.at[index, "Email Name"]            = class_id
            continue

        data_df.at[index, "D4 Code"] = dim_row.iloc[0, 2]
        data_df.at[index, "D4 Name"] = dim_row.iloc[0, 3]

        if grouping_type == "Centre":
            data_df.at[index, "Invoice Grouping Code"] = dim_row.iloc[0, 0]
            data_df.at[index, "Invoice Grouping Name"] = dim_row.iloc[0, 1]
            data_df.at[index, "Invoice Customer ID"]   = dim_row.iloc[0, 0]
        elif grouping_type == "Site":
            data_df.at[index, "Invoice Grouping Code"] = dim_row.iloc[0, 4]
            data_df.at[index, "Invoice Grouping Name"] = dim_row.iloc[0, 5]
            data_df.at[index, "Invoice Customer ID"]   = dim_row.iloc[0, 4]
        elif grouping_type == "Group":
            data_df.at[index, "Invoice Grouping Code"] = dim_row.iloc[0, 6]
            data_df.at[index, "Invoice Grouping Name"] = dim_row.iloc[0, 7]
            data_df.at[index, "Invoice Customer ID"]   = dim_row.iloc[0, 6]
        elif grouping_type == "Custom":
            data_df.at[index, "Invoice Grouping Code"] = dim_row.iloc[0, 16]
            data_df.at[index, "Invoice Grouping Name"] = dim_row.iloc[0, 17]
            data_df.at[index, "Invoice Customer ID"]   = dim_row.iloc[0, 16]
        else:
            print(f"⚠️ Unrecognised Invoice_Grouping_Type '{grouping_type}' "
                  f"for Class ID '{class_id}'. Falling back to Class ID.")
            data_df.at[index, "Invoice Grouping Code"] = class_id
            data_df.at[index, "Invoice Grouping Name"] = class_id
            data_df.at[index, "Invoice Customer ID"]   = class_id

        email_type = row["Email_Grouping_Type"]
        if email_type == "Centre":
            data_df.at[index, "Email Code"] = dim_row.iloc[0, 0]
            data_df.at[index, "Email Name"] = dim_row.iloc[0, 1]
        elif email_type == "Site":
            data_df.at[index, "Email Code"] = dim_row.iloc[0, 4]
            data_df.at[index, "Email Name"] = dim_row.iloc[0, 5]
        elif email_type == "Group":
            data_df.at[index, "Email Code"] = dim_row.iloc[0, 6]
            data_df.at[index, "Email Name"] = dim_row.iloc[0, 7]
        else:
            data_df.at[index, "Email Code"] = class_id
            data_df.at[index, "Email Name"] = class_id

    return data_df


# Financial year period-to-month mapping (post-extension, 2026 onward).
# The 2026 FY was extended to run through December: Aug-Dec now continue as
# periods 13-17 instead of wrapping back to period 1. Periods 8-12 are
# no longer used. Old-format codes (pre-extension) are historical only and
# are not reprocessed, so no dual-scheme handling is needed here.
PERIOD_MONTH_ABBR = {
    1: "Jan", 2: "Feb", 3: "Mar", 4: "Apr", 5: "May", 6: "Jun", 7: "Jul",
    13: "Aug", 14: "Sep", 15: "Oct", 16: "Nov", 17: "Dec",
}
PERIOD_MONTH_LONG = {
    1: "January", 2: "February", 3: "March", 4: "April", 5: "May", 6: "June", 7: "July",
    13: "August", 14: "September", 15: "October", 16: "November", 17: "December",
}


def _parse_period_code(period_code):
    """Split a period code like '2026/013' or bare '13' into (year_part, period_num).
    period_num is normalised to an int so it doesn't matter whether the source
    system sends '013', '13', '01', etc. Returns (year_part, None) if the
    period number can't be parsed.
    """
    code_str = str(period_code).strip()
    if not code_str or code_str.lower() in ("nan", "unknown", "none", ""):
        return None, None

    parts = code_str.split('/')
    if len(parts) > 1:
        year_part, period_part = parts[0], parts[-1]
    else:
        year_part, period_part = None, parts[0]

    try:
        period_num = int(period_part)
    except (ValueError, TypeError):
        return year_part, None

    return year_part, period_num


def _build_period_display(period_code, posting_date):
    """Convert a period code (e.g. '2026/013') to a short display string (e.g. 'Aug 2026')."""
    year_part, period_num = _parse_period_code(period_code)
    month_abbr = PERIOD_MONTH_ABBR.get(period_num, "")

    if not month_abbr:
        print(f"⚠️ _build_period_display: period number '{period_num}' (from '{period_code}') not in lookup")
        return posting_date.strftime("%b %Y")

    year = year_part if year_part else str(posting_date.year)
    return f"{month_abbr} {year}"


def _build_period_display_long(period_code):
    """Convert a period code to a 'P{n} Month' label for the A1 client label,
    sheet tab, and filename (e.g. 'P13 August').

    Handles full period codes like '2026/013' and bare numbers like '13'.
    Falls back gracefully — never returns 'Unknown', instead logs a warning and
    returns an empty string so the caller can decide what to show.
    """
    year_part, period_num = _parse_period_code(period_code)
    if period_num is None:
        print(f"⚠️ _build_period_display_long: empty/unparseable period code '{period_code}'")
        return ""

    month_name = PERIOD_MONTH_LONG.get(period_num)
    if not month_name:
        print(f"⚠️ _build_period_display_long: period number '{period_num}' not in lookup (from '{period_code}')")
        return ""

    print(f"✅ Period code '{period_code}' -> P{period_num} {month_name}")
    return f"P{period_num} {month_name}"


def generate_summary_sheets(data_df, units_df, mark_up_adjustments_df, dimensions_df):
    summaries = {}
    for class_id, class_group in data_df.groupby("Class ID"):
        summary_data = []
        unit_code = class_group["Unit Code"].iloc[0]
        unit_row = units_df[units_df.iloc[:, 0] == unit_code]
        summary_format = unit_row.iloc[0, 17] if not unit_row.empty else "Standard"

        # Accumulators for the three summary header rows
        total_supplier_net, total_supplier_tax, total_supplier_gross = 0.0, 0.0, 0.0
        total_markup_net, total_markup_tax, total_markup_gross = 0.0, 0.0, 0.0

        for supplier_id, supplier_group in class_group.groupby("Supplier ID"):
            supplier_total_net, supplier_total_tax, supplier_total_gross = 0, 0, 0
            for _, row in supplier_group.iterrows():
                markup = get_markup_rate(row, mark_up_adjustments_df)
                net, tax, gross = row["Transaction Amount"], row["Transaction Tax"], row["Total Amount"]
                markup_gross = gross * markup if markup else 0
                if markup and tax != 0 and gross != 0:
                    markup_tax = markup_gross * (tax / gross)
                else:
                    markup_tax = 0.0
                markup_net = markup_gross - markup_tax

                summary_data.append({
                    "Supplier": row["Supplier Name"],
                    "Invoice Number": row["AP Purchase Invoice Number"],
                    "Date": row["Date"].strftime("%d/%m/%Y") if pd.notna(row["Date"]) else "",
                    "Center": row["Invoice Grouping Name"],
                    "Net (£)": net,
                    "VAT (£)": tax,
                    "Gross (£)": round(gross, 2),
                })
                supplier_total_net   += net
                supplier_total_tax   += tax
                supplier_total_gross += gross

                total_markup_net   += markup_net
                total_markup_tax   += markup_tax
                total_markup_gross += markup_gross

            summary_data.append({
                "Supplier": f"{row['Supplier Name']} Subtotal",
                "Invoice Number": "",
                "Date": "",
                "Center": "",
                "Net (£)": round(supplier_total_net, 2),
                "VAT (£)": round(supplier_total_tax, 2),
                "Gross (£)": round(supplier_total_net + supplier_total_tax, 2),
            })
            total_supplier_net   += supplier_total_net
            total_supplier_tax   += supplier_total_tax
            total_supplier_gross += supplier_total_gross

        summary_df = pd.DataFrame(summary_data)

        grand_net   = round(total_supplier_net   + total_markup_net,   2)
        grand_tax   = round(total_supplier_tax   + total_markup_tax,   2)
        grand_gross = round(total_supplier_gross + total_markup_gross, 2)

        totals_block = {
            "supplier_total": {
                "Net (£)":   round(total_supplier_net,   2),
                "VAT (£)":   round(total_supplier_tax,   2),
                "Gross (£)": round(total_supplier_gross, 2),
            },
            "markup_total": {
                "Net (£)":   round(total_markup_net,   2),
                "VAT (£)":   round(total_markup_tax,   2),
                "Gross (£)": round(total_markup_gross, 2),
            },
            "grand_total": {
                "Net (£)":   grand_net,
                "VAT (£)":   grand_tax,
                "Gross (£)": grand_gross,
            },
        }

        client_name = get_client_name(class_id, dimensions_df)
        period_code = class_group["Period Code"].iloc[0]
        print(f"🔍 Period code raw value for {class_id}: '{period_code}' (type: {type(period_code).__name__})")
        month_name  = _build_period_display_long(period_code)

        summaries[class_id] = (summary_df, totals_block, client_name, month_name)
    return summaries


def get_markup_rate(row, mark_up_adjustments_df):
    mark_up_check = row["Mark Up Check"]
    adjustment_row = mark_up_adjustments_df[mark_up_adjustments_df.iloc[:, 0] == mark_up_check]
    if not adjustment_row.empty:
        return adjustment_row.iloc[0, 1]
    return row["Mark Up"] if pd.notna(row["Mark Up"]) else 0


def get_payment_terms(unit_code, billing_matrix_df):
    """Look up payment terms (days) from column O of the billing matrix using Unit Code in column A."""
    parts = unit_code.split('/')
    candidates = []

    candidates.append(unit_code)

    for i in range(len(parts) - 1, 0, -1):
        prefix = '/'.join(parts[:i])
        if prefix != unit_code:
            candidates.append(prefix)

    col_a = billing_matrix_df.iloc[:, 0].astype(str)

    for candidate in candidates:
        billing_row = billing_matrix_df[col_a == candidate]
        if not billing_row.empty:
            payment_terms_value = billing_row.iloc[0, 14]  # Column O = index 14
            if pd.notna(payment_terms_value) and payment_terms_value != '':
                try:
                    days = int(payment_terms_value)
                    print(f"📅 Payment terms for '{candidate}' (from unit code '{unit_code}'): {days} days")
                    return days
                except (ValueError, TypeError):
                    print(f"⚠️ Invalid payment terms value '{payment_terms_value}' for '{candidate}', trying next candidate")

    print(f"⚠️ No payment terms found for unit code '{unit_code}' (tried: {candidates}), using default 30 days")
    return 30


def create_invoice_template(data_df, units_df, mark_up_adjustments_df, processed_date):
    invoice_lines = []

    billing_matrix_df = load_from_s3(CONFIG_BUCKET, "Client Billing Matrix - Sage Version.xlsx", "Units")

    valid_data_df = data_df[
        data_df["Invoice Grouping Code"].notna() &
        (data_df["Invoice Grouping Code"].astype(str).str.strip() != "") &
        (data_df["Invoice Grouping Code"].astype(str).str.strip() != "Unknown")
    ]
    if len(valid_data_df) < len(data_df):
        dropped = len(data_df) - len(valid_data_df)
        print(f"⚠️ Skipped {dropped} rows with missing Invoice Grouping Code in invoice template")

    for grouping_id, group in valid_data_df.groupby("Invoice Grouping Code"):

        customer_id = group["Invoice Customer ID"].iloc[0] if "Invoice Customer ID" in group.columns else group["Customer ID"].iloc[0]
        if "Invoice Customer ID" not in group.columns or pd.isna(customer_id) or str(customer_id).strip() in ("", "Unknown"):
            customer_id = group["Customer ID"].iloc[0]
            print(f"⚠️ Invoice Customer ID missing for group {grouping_id}, falling back to source Customer ID: {customer_id}")
        else:
            print(f"📋 Billing group {grouping_id} to customer: {customer_id}")

        min_date = group["Date"].min()
        if pd.isna(min_date):
            posting_date = eom_date(datetime.now())
            print(f"⚠️ Warning: No valid dates for group {grouping_id}, using current date")
        else:
            posting_date = eom_date(min_date)

        created_date = posting_date

        unit_code_for_lookup = group["Unit Code"].iloc[0] if "Unit Code" in group.columns else None
        payment_terms_days   = get_payment_terms(unit_code_for_lookup, billing_matrix_df) if unit_code_for_lookup else 30
        due_date             = processed_date + timedelta(days=payment_terms_days)

        print(f"📅 Group '{grouping_id}': posting={posting_date.strftime('%d/%m/%Y') if posting_date else 'N/A'}, "
              f"processed={processed_date.strftime('%d/%m/%Y')}, "
              f"due={due_date.strftime('%d/%m/%Y')} ({payment_terms_days} days)")

        posting_date_str = posting_date.strftime("%d/%m/%Y") if posting_date else ""
        created_date_str = created_date.strftime("%d/%m/%Y") if created_date else ""
        due_date_str     = due_date.strftime("%d/%m/%Y")

        dimension_type = determine_dimension_type(group["Unit Code"].iloc[0], units_df)
        gl_code        = get_gl_code(dimension_type)
        grouping_type  = group["Invoice_Grouping_Type"].iloc[0] if "Invoice_Grouping_Type" in group.columns else None

        if grouping_type == "Group":

            d2_name          = group["Invoice Grouping Name"].iloc[0] if "Invoice Grouping Name" in group.columns else grouping_id
            period_code      = group["Period Code"].iloc[0] if "Period Code" in group.columns else None
            period_display   = _build_period_display(period_code, posting_date)
            description_text = f"{d2_name} Purchases for {period_display}"

            full_total_gross = 0.0
            for _, row in group.iterrows():
                markup_rate = get_markup_rate(row, mark_up_adjustments_df)
                row_gross   = row["Total Amount"]
                row_tax     = row["Transaction Tax"]
                if markup_rate:
                    markup_gross  = row_gross * markup_rate
                    markup_tax    = markup_gross * (row_tax / row_gross) if row_tax != 0 and row_gross != 0 else 0.0
                    row_gross    += markup_gross
                full_total_gross += row_gross

            line_no    = 1
            first_line = True

            for d4_code, d4_group in group.groupby("D4 Code"):

                d4_total_gross = 0.0
                d4_total_tax   = 0.0
                for _, row in d4_group.iterrows():
                    markup_rate = get_markup_rate(row, mark_up_adjustments_df)
                    row_gross   = row["Total Amount"]
                    row_tax     = row["Transaction Tax"]
                    if markup_rate:
                        markup_gross  = row_gross * markup_rate
                        markup_tax    = markup_gross * (row_tax / row_gross) if row_tax != 0 and row_gross != 0 else 0.0
                        row_gross    += markup_gross
                        row_tax      += markup_tax
                    d4_total_gross += row_gross
                    d4_total_tax   += row_tax

                d4_net_0, d4_net_20, d4_tax_20 = calculate_tax_split(d4_total_gross, d4_total_tax)

                d4_name   = d4_group["D4 Name"].iloc[0] if "D4 Name" in d4_group.columns else d4_code
                memo_text = f"{d4_name} Purchases for {period_display}"

                print(f"📋 Group '{grouping_id}' — D4 line {line_no}: {d4_code} ({d4_name}), gross: {round(d4_total_gross, 2)}")

                if d4_net_0 != 0:
                    invoice_lines.append({
                        "DONOTIMPORT":              grouping_id if first_line else "",
                        "INVOICE_NO":               "",
                        "PO_NO":                    "",
                        "CUSTOMER_ID":              customer_id if first_line else "",
                        "posting_date":             posting_date_str if first_line else "",
                        "CREATED_DATE":             created_date_str if first_line else "",
                        "due_date":                 due_date_str if first_line else "",
                        "TOTAL_DUE":                round(full_total_gross, 2) if first_line else "",
                        "Description":              description_text if first_line else "",
                        "LINE_NO":                  line_no,
                        "MEMO":                     memo_text,
                        "ACCT_NO":                  gl_code,
                        "LOCATION_ID":              "AMH",
                        "AMOUNT":                   d4_net_0,
                        "SUPDOCID":                 "",
                        "TAX_LINE_NO":              1,
                        "TAX_AMOUNT":               0,
                        "TAX_DETAILID":             "UK Sale Goods Zero Rate",
                        "ARINVOICEITEM_CLASSID":    d4_code,
                        "ARINVOICEITEM_CUSTOMERID": customer_id
                    })
                    first_line = False
                    line_no += 1

                if d4_net_20 != 0:
                    invoice_lines.append({
                        "DONOTIMPORT":              grouping_id if first_line else "",
                        "INVOICE_NO":               "",
                        "PO_NO":                    "",
                        "CUSTOMER_ID":              customer_id if first_line else "",
                        "posting_date":             posting_date_str if first_line else "",
                        "CREATED_DATE":             created_date_str if first_line else "",
                        "due_date":                 due_date_str if first_line else "",
                        "TOTAL_DUE":                round(full_total_gross, 2) if first_line else "",
                        "Description":              description_text if first_line else "",
                        "LINE_NO":                  line_no,
                        "MEMO":                     memo_text,
                        "ACCT_NO":                  gl_code,
                        "LOCATION_ID":              "AMH",
                        "AMOUNT":                   d4_net_20,
                        "SUPDOCID":                 "",
                        "TAX_LINE_NO":              1,
                        "TAX_AMOUNT":               d4_tax_20,
                        "TAX_DETAILID":             "UK Sale Goods Standard Rate",
                        "ARINVOICEITEM_CLASSID":    d4_code,
                        "ARINVOICEITEM_CUSTOMERID": customer_id
                    })
                    first_line = False
                    line_no += 1

        else:
            base_total_gross = group["Total Amount"].sum()
            base_total_tax   = group["Transaction Tax"].sum()
            total_markup     = 0
            markup_tax       = 0
            for _, row in group.iterrows():
                markup_rate = get_markup_rate(row, mark_up_adjustments_df)
                if markup_rate:
                    row_markup_gross = row["Total Amount"] * markup_rate
                    total_markup += row_markup_gross
                    if row["Transaction Tax"] != 0 and row["Total Amount"] != 0:
                        effective_tax_rate = row["Transaction Tax"] / row["Total Amount"]
                        markup_tax += row_markup_gross * effective_tax_rate
            total_gross = base_total_gross + total_markup
            total_tax   = base_total_tax + markup_tax

            grouping_name  = group["Invoice Grouping Name"].iloc[0] if "Invoice Grouping Name" in group.columns else grouping_id
            period_code    = group["Period Code"].iloc[0] if "Period Code" in group.columns else None
            period_display = _build_period_display(period_code, posting_date)
            memo_text      = f"{grouping_name} Purchases for {period_display}"

            net_0, net_20, tax_20 = calculate_tax_split(total_gross, total_tax)

            # BUG FIX: previously both the net_0 and net_20 blocks hardcoded
            # "LINE_NO": 1, and a third block then re-appended net_20 again at
            # LINE_NO 2. On mixed-rate invoices (both net_0 and net_20 non-zero)
            # this wrote the standard-rate amount TWICE — once on a duplicate
            # LINE_NO 1 row, once on LINE_NO 2 — inflating AMOUNT/TAX_AMOUNT
            # and TOTAL_DUE by the standard-rate value. Fixed by using a single
            # running line_no counter and dropping the redundant third block.
            line_no = 1

            if net_0 != 0:
                invoice_lines.append({
                    "DONOTIMPORT":              grouping_id,
                    "INVOICE_NO":               "",
                    "PO_NO":                    "",
                    "CUSTOMER_ID":              customer_id,
                    "posting_date":             posting_date_str,
                    "CREATED_DATE":             created_date_str,
                    "due_date":                 due_date_str,
                    "TOTAL_DUE":                round(total_gross, 2),
                    "Description":              memo_text,
                    "LINE_NO":                  line_no,
                    "MEMO":                     memo_text,
                    "ACCT_NO":                  gl_code,
                    "LOCATION_ID":              "AMH",
                    "AMOUNT":                   net_0,
                    "SUPDOCID":                 "",
                    "TAX_LINE_NO":              1,
                    "TAX_AMOUNT":               0,
                    "TAX_DETAILID":             "UK Sale Goods Zero Rate",
                    "ARINVOICEITEM_CLASSID":    group["D4 Code"].iloc[0],
                    "ARINVOICEITEM_CUSTOMERID": customer_id
                })
                line_no += 1

            if net_20 != 0:
                is_first_line = (line_no == 1)
                invoice_lines.append({
                    "DONOTIMPORT":              grouping_id if is_first_line else "",
                    "INVOICE_NO":               "",
                    "PO_NO":                    "",
                    "CUSTOMER_ID":              customer_id if is_first_line else "",
                    "posting_date":             posting_date_str if is_first_line else "",
                    "CREATED_DATE":             created_date_str if is_first_line else "",
                    "due_date":                 due_date_str if is_first_line else "",
                    "TOTAL_DUE":                round(total_gross, 2) if is_first_line else "",
                    "Description":              memo_text if is_first_line else "",
                    "LINE_NO":                  line_no,
                    "MEMO":                     memo_text,
                    "ACCT_NO":                  gl_code,
                    "LOCATION_ID":              "AMH",
                    "AMOUNT":                   net_20,
                    "SUPDOCID":                 "",
                    "TAX_LINE_NO":              1,
                    "TAX_AMOUNT":               tax_20,
                    "TAX_DETAILID":             "UK Sale Goods Standard Rate",
                    "ARINVOICEITEM_CLASSID":    group["D4 Code"].iloc[0],
                    "ARINVOICEITEM_CUSTOMERID": customer_id
                })

    return pd.DataFrame(invoice_lines)


def determine_dimension_type(unit_code, units_df):
    """Determine dimension type by exact unit code match in the billing matrix (column A),
    then read the unit name (column B). Never uses prefix/substring matching on the code."""
    unit_row = units_df[units_df.iloc[:, 0].astype(str) == str(unit_code)]
    if not unit_row.empty:
        unit_name = str(unit_row.iloc[0, 1])
        if "Education" in unit_name:
            return "Education"
        elif "Care" in unit_name:
            return "Care"
        elif "Leisure" in unit_name:
            return "Leisure"
        elif "Contractor" in unit_name:
            return "Contractor"
        else:
            print(f"⚠️ Unit '{unit_code}' has unrecognised name '{unit_name}', defaulting to Education")
    else:
        print(f"⚠️ Unit code '{unit_code}' not found in billing matrix, defaulting to Education")
    return "Education"


def create_email_structure(data_df, summaries, output_bucket, dimensions_df):
    email_structure = []
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    for class_id, (summary_df, totals_block, client_name, month_name) in summaries.items():
        unit_code = data_df[data_df["Class ID"] == class_id]["Unit Code"].iloc[0]

        clean_name   = clean_client_label(client_name)
        # Only append the period separator if month_name resolved to something
        if month_name:
            sheet_tab    = f"{clean_name} {month_name}"[:31]
            client_label = f"{clean_name} - {month_name}"
        else:
            sheet_tab    = clean_name[:31]
            client_label = clean_name

        file_name = f"summaries/{class_id.replace('/', '-')}_INV_SUMM_{clean_name}_{month_name}_{timestamp}.xlsx"
        upload_summary_to_s3(
            summary_df, output_bucket, file_name, totals_block,
            sheet_name=sheet_tab,
            client_label=client_label
        )
        email_structure.append({
            "Unit Code": unit_code,
            "Document Grouping": class_id,
            "Email Code": data_df[data_df["Class ID"] == class_id]["Email Code"].iloc[0],
            "File Path": f"s3://{output_bucket}/{file_name}",
            "Doc.Type": "I-Summary",
            "Missing Details": ""
        })
    return pd.DataFrame(email_structure)


def lambda_handler(event, context):
    try:
        print("🚀 === LAMBDA START ===")

        processed_date = datetime.now()
        print(f"🕐 Processing timestamp: {processed_date.strftime('%d/%m/%Y %H:%M:%S')}")

        bucket = None
        key    = None

        if 'Records' in event and len(event['Records']) > 0:
            bucket = event['Records'][0]['s3']['bucket']['name']
            key    = urllib.parse.unquote_plus(event['Records'][0]['s3']['object']['key'])
            print(f"📂 S3 Trigger - Bucket: {bucket}, Key: {key}")
        elif 'bucket' in event and 'key' in event:
            bucket = event['bucket']
            key    = event['key']
            print(f"📂 Manual Trigger - Bucket: {bucket}, Key: {key}")
        else:
            raise ValueError("❌ Could not determine bucket and key from event")

        print(f"🔍 Processing file: s3://{bucket}/{key}")

        file_id             = None
        can_update_dynamodb = False

        print("🔍 Step 1: Trying to get file ID from S3 metadata...")
        file_id = get_file_id_from_s3_metadata(bucket, key)

        if not file_id:
            print("🔍 Step 2: Trying to find file ID by S3 key...")
            file_id = find_file_id_by_s3_key(key)

        if not file_id:
            print("❌ CRITICAL: Could not find file ID in DynamoDB")
            print("💡 This usually means:")
            print("   - FileUpload didn't save the s3Key properly to DynamoDB")
            print("   - The s3Key in DynamoDB doesn't match the actual S3 key")
            print("   - FILE_METADATA_TABLE environment variable is wrong")
            can_update_dynamodb = False
        else:
            can_update_dynamodb = True
            print(f"✅ Found file ID: {file_id}")

        if can_update_dynamodb:
            print("🔄 Updating status to PROCESSING...")
            update_result = update_file_metadata(file_id, {
                'status': 'PROCESSING',
                'processedDate': processed_date.isoformat()
            })
            if not update_result:
                print("❌ Failed to update DynamoDB, continuing without updates")
                can_update_dynamodb = False

        print("📊 Loading and processing file...")
        report_df = load_from_s3(bucket, key)

        print("📂 Loading configuration files...")
        units_df               = load_from_s3(CONFIG_BUCKET, "Client Billing Matrix - Sage Version.xlsx", "Units")
        dimensions_df          = load_from_s3(CONFIG_BUCKET, "Client Dimensions - Sage version.xlsx", "Dimensions")
        mark_up_adjustments_df = load_from_s3(CONFIG_BUCKET, "Client Billing Matrix - Sage Version.xlsx", "Mark Up Adjustments")

        print("⚙️ Processing billing data...")
        report_df          = validate_units(report_df, units_df)
        data_df            = create_data_sheet(report_df, units_df, dimensions_df)
        summaries          = generate_summary_sheets(data_df, units_df, mark_up_adjustments_df, dimensions_df)
        invoice_df         = create_invoice_template(data_df, units_df, mark_up_adjustments_df, processed_date)
        email_structure_df = create_email_structure(data_df, summaries, OUTPUT_BUCKET, dimensions_df)

        timestamp       = processed_date.strftime("%Y%m%d_%H%M%S")
        input_file_name = key.split('/')[-1].replace('.csv', '').replace('.xlsx', '').replace(' ', '_')

        output_files = []

        data_file_key = f"outputs/{input_file_name}_Data_{timestamp}.xlsx"
        upload_to_s3(data_df, OUTPUT_BUCKET, data_file_key)
        output_files.append({
            'type': 'DATA_SHEET',
            's3Key': data_file_key,
            'fileName': f"{input_file_name}_Data_{timestamp}.xlsx",
            'category': 'MAIN_OUTPUT'
        })

        invoice_file_key = f"outputs/{input_file_name}_Invoice_Template_{timestamp}.csv"
        upload_to_s3(invoice_df, OUTPUT_BUCKET, invoice_file_key, 'csv')
        output_files.append({
            'type': 'INVOICE_TEMPLATE',
            's3Key': invoice_file_key,
            'fileName': f"{input_file_name}_Invoice_Template_{timestamp}.csv",
            'category': 'MAIN_OUTPUT'
        })

        email_file_key = f"outputs/{input_file_name}_Email_Structure_{timestamp}.xlsx"
        upload_to_s3(email_structure_df, OUTPUT_BUCKET, email_file_key)
        output_files.append({
            'type': 'EMAIL_STRUCTURE',
            's3Key': email_file_key,
            'fileName': f"{input_file_name}_Email_Structure_{timestamp}.xlsx",
            'category': 'MAIN_OUTPUT'
        })

        for class_id, (summary_df, totals_block, client_name, month_name) in summaries.items():
            safe_class_id = class_id.replace("/", "-")
            clean_name    = clean_client_label(client_name)
            safe_client   = clean_name.replace(" ", "_").replace("/", "-")

            # Only include period separator in tab/label/filename if month resolved
            if month_name:
                sheet_tab    = f"{clean_name} {month_name}"[:31]
                client_label = f"{clean_name} - {month_name}"
                safe_month   = month_name.replace(" ", "_")
            else:
                sheet_tab    = clean_name[:31]
                client_label = clean_name
                safe_month   = "UnknownPeriod"

            summary_file_key = f"summaries/{input_file_name}_Summary_{safe_client}_{safe_class_id}_{timestamp}.xlsx"
            upload_summary_to_s3(
                summary_df, OUTPUT_BUCKET, summary_file_key, totals_block,
                sheet_name=sheet_tab,
                client_label=client_label
            )
            output_files.append({
                'type': 'CLIENT_SUMMARY',
                's3Key': summary_file_key,
                'fileName': f"{input_file_name}_Summary_{safe_client}_{safe_class_id}.xlsx",
                'category': 'SUMMARY',
                'clientName': clean_name,
                'classId': class_id
            })

        print("✅ Billing processing completed successfully!")

        if can_update_dynamodb:
            print("🔗 Generating presigned download URLs...")
            download_urls = generate_presigned_urls(output_files)

            print("💾 Updating DynamoDB with output files and download URLs...")
            update_result = update_file_metadata(file_id, {
                'status': 'PROCESSED',
                'outputFiles': output_files,
                'downloadUrls': download_urls,
                'processedDate': processed_date.isoformat()
            })

            if update_result:
                print("🎉 SUCCESS: DynamoDB updated with output files and download URLs!")
                print(f"📥 {len(download_urls)} files ready for auto-download")
            else:
                print("❌ WARNING: Failed to update DynamoDB with output files")
        else:
            print("⚠️  SKIPPED: Could not update DynamoDB (no file ID found)")

        print("🏁 === LAMBDA COMPLETED ===")

        return {
            'status': 'SUCCESS',
            'inputFile': key,
            'outputFiles': output_files,
            'processedRecords': len(data_df),
            'summaryCount': len(summaries),
            'timestamp': timestamp,
            'processedDate': processed_date.isoformat(),
            'dynamodbUpdated': can_update_dynamodb,
            'downloadUrlsGenerated': can_update_dynamodb
        }

    except Exception as e:
        error_msg = f"❌ Error processing billing data: {str(e)}\n{traceback.format_exc()}"
        print(error_msg)

        if 'file_id' in locals() and file_id and 'can_update_dynamodb' in locals() and can_update_dynamodb:
            print("🔄 Updating status to FAILED in DynamoDB...")
            update_file_metadata(file_id, {
                'status': 'FAILED',
                'errorMessage': str(e)
            })

        return {
            'status': 'FAILED',
            'error': str(e)
        }